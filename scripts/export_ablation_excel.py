"""Export ablation results to Excel."""
import json, os, glob, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

base = "results_v2/ablation"
rows = []

for d in sorted(os.listdir(base)):
    dpath = os.path.join(base, d)
    if not os.path.isdir(dpath):
        continue

    summary_files = glob.glob(os.path.join(dpath, "*_summary.json"))
    if not summary_files:
        continue

    for sf in summary_files:
        if sf.endswith("/summary.json"):
            continue
        with open(sf) as f:
            s = json.load(f)

        tag = s.get("tag", d)
        mode = s.get("mode", "")
        strategy = s.get("strategy", "")
        accuracy = s.get("accuracy", 0)
        correct = s.get("correct", 0)
        total = s.get("total", 100)
        done = s.get("done", True)
        avg_fwd = s.get("avg_forward_passes", 0)
        avg_output_tokens = s.get("avg_output_tokens", 0)
        time_s = s.get("time_s", 0)
        editing_threshold = s.get("editing_threshold", "")

        if mode == "original":
            tau = ""
            c_max = ""
            rho = ""
            avg_token_mods = s.get("avg_t2t_edits", 0)
            mod_type = "T2T edits"
        else:
            tau = s.get("remask_threshold", "")
            c_max_match = re.search(r"_c(\d+)_", tag)
            c_max = int(c_max_match.group(1)) if c_max_match else ""
            rho = s.get("max_remask_ratio", "")
            avg_token_mods = s.get("avg_remask_total", 0)
            mod_type = "remasks"

        rows.append({
            "dir": d,
            "strategy": strategy if strategy else "original",
            "tau": tau,
            "c_max": c_max,
            "rho": rho,
            "accuracy_pct": round(accuracy * 100, 1),
            "correct": correct,
            "total": total,
            "done": done,
            "editing_threshold": editing_threshold,
            "avg_token_modifications": round(avg_token_mods, 1),
            "modification_type": mod_type,
            "avg_forward_passes": round(avg_fwd, 1),
            "avg_output_tokens": round(avg_output_tokens, 1),
            "time_s": round(time_s, 1),
        })

wb = Workbook()
ws = wb.active
ws.title = "Ablation Results (CMATH)"

headers = [
    "Config ID", "Strategy", "τ (threshold)", "C (max cycles)", "ρ (remask ratio)",
    "Accuracy (%)", "Correct", "Total", "Done",
    "τ_edit", "Avg Token Modifications", "Modification Type",
    "Avg Forward Passes", "Avg Output Tokens", "Time (s)"
]

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

strategy_fills = {
    "original": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
    "low_prob": PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid"),
    "t2t_remask": PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid"),
    "logit_diff": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
}

best_acc = max(r["accuracy_pct"] for r in rows)
best_font = Font(bold=True, color="FF0000")

for row_idx, r in enumerate(rows, 2):
    values = [
        r["dir"], r["strategy"], r["tau"], r["c_max"], r["rho"],
        r["accuracy_pct"], r["correct"], r["total"],
        "Yes" if r["done"] else "No",
        r["editing_threshold"],
        r["avg_token_modifications"], r["modification_type"],
        r["avg_forward_passes"], r["avg_output_tokens"], r["time_s"]
    ]
    fill = strategy_fills.get(r["strategy"], PatternFill())
    for col_idx, v in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=v)
        cell.border = thin_border
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
    if r["accuracy_pct"] == best_acc:
        ws.cell(row=row_idx, column=6).font = best_font

col_widths = [35, 14, 14, 16, 16, 12, 8, 6, 6, 8, 22, 16, 18, 18, 10]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"
ws.freeze_panes = "A2"

out_path = "paper/ablation_tradeoff_results.xlsx"
wb.save(out_path)
print(f"Saved {len(rows)} rows to {out_path}")
